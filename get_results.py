from openpyxl import load_workbook
workbook = load_workbook(filename="TdF Stages Riders Results (2011-).xlsx", read_only=True, data_only=True)

def year_results(year):
    ws = workbook[year]
    res_md = f"## {ws['A1'].value}\n"
    for i in range(2, ws.max_row+1):
        r = ws[i]
        md = "| "
        for c in r:
            md += f'{c.value} | '
        md += '\n'
        if i == 2:
            md += '| --- | --- | --- | --- | --- | --- |\n'
        #print(md)
        res_md += md
    return res_md

with open('results_file.md', 'w') as resfile:
    for y in range(12, 24):
        #print(year_results(f'TdF20{str(y)} Results'))
        resfile.write(year_results(f'TdF20{str(y)} Results'))
        resfile.write('\n\n')
