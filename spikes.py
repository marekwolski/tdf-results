from openpyxl import load_workbook
workbook = load_workbook(filename="TdF Stages Riders Results (2011-).xlsx", read_only=True)
#print(workbook.sheetnames)
#exit()

#results = workbook['TdF2013 Results']
#for row in results:
#    print(row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value)
#    for cell in row:
#        print(cell.value)

def year_results(year):
    ws = workbook[year]
    res_md = f"## {ws['A1'].value}\n"
    for i in range(2, ws.max_row+1):
        r = ws[i]
        md = "| "
        for c in r:
            md += f'{c.value} | '
        md += '\n'
        if i == 2:
            md += '| --- | --- | --- | --- | --- | --- |\n'
        #print(md)
        res_md += md
    return res_md

with open('results_file.md', 'w') as resfile:
    for y in range(12, 24):
        #print(year_results(f'TdF20{str(y)} Results'))
        resfile.write(year_results(f'TdF20{str(y)} Results'))
        resfile.write('\n\n')
